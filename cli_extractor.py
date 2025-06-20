#!/usr/bin/env python3
"""
명령줄 웹 텍스트 추출 및 번역 도구
간단한 사용법: python cli_extractor.py <URL> [출력파일명]
"""

import sys
import argparse
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from googletrans import Translator
import re
from datetime import datetime

class CLIWebTextExtractor:
    def __init__(self):
        self.translator = Translator()
    
    def extract_text_from_url(self, url, verbose=True):
        """웹페이지에서 텍스트 추출"""
        try:
            if verbose:
                print(f"웹페이지 접속 중: {url}")
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            response.encoding = response.apparent_encoding
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # 불필요한 태그 제거
            for tag in soup(['script', 'style', 'nav', 'header', 'footer', 'aside']):
                tag.decompose()
            
            text_elements = []
            
            # 제목들 추출
            for heading in soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6']):
                text = heading.get_text(strip=True)
                if text and len(text) > 1:
                    text_elements.append({
                        'type': 'heading',
                        'tag': heading.name,
                        'text': text
                    })
            
            # 본문 텍스트 추출
            for para in soup.find_all(['p', 'div', 'span', 'li']):
                text = para.get_text(strip=True)
                if text and len(text) > 10:
                    # 중복 제거
                    is_duplicate = False
                    for existing in text_elements:
                        if text in existing['text'] or existing['text'] in text:
                            is_duplicate = True
                            break
                    
                    if not is_duplicate:
                        text_elements.append({
                            'type': 'content',
                            'tag': para.name,
                            'text': text
                        })
            
            if verbose:
                print(f"총 {len(text_elements)}개의 텍스트 요소를 추출했습니다.")
            
            return text_elements
            
        except Exception as e:
            print(f"텍스트 추출 오류: {str(e)}")
            return []
    
    def translate_text(self, text, target_lang, verbose=True):
        """텍스트 번역"""
        try:
            if len(text) > 4000:
                # 긴 텍스트 분할 처리
                sentences = re.split(r'[.!?。！？]', text)
                translated_sentences = []
                
                current_chunk = ""
                for sentence in sentences:
                    if len(current_chunk + sentence) < 4000:
                        current_chunk += sentence + "."
                    else:
                        if current_chunk:
                            result = self.translator.translate(current_chunk, dest=target_lang)
                            translated_sentences.append(result.text)
                        current_chunk = sentence + "."
                
                if current_chunk:
                    result = self.translator.translate(current_chunk, dest=target_lang)
                    translated_sentences.append(result.text)
                
                return " ".join(translated_sentences)
            else:
                result = self.translator.translate(text, dest=target_lang)
                return result.text
                
        except Exception as e:
            if verbose:
                print(f"번역 오류 ({target_lang}): {str(e)}")
            return f"[번역 실패: {text[:50]}...]"
    
    def create_excel_file(self, text_elements, file_path, languages=['en', 'zh-cn', 'vi'], verbose=True):
        """엑셀 파일 생성"""
        try:
            if verbose:
                print("엑셀 파일 생성 중...")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "웹 텍스트 추출 결과"
            
            # 헤더 설정
            headers = ["번호", "유형", "태그", "원본 텍스트(한국어)"]
            
            lang_names = {
                'en': '영어',
                'zh-cn': '중국어',
                'vi': '베트남어',
                'ja': '일본어',
                'es': '스페인어',
                'fr': '프랑스어'
            }
            
            for lang in languages:
                headers.append(f"{lang_names.get(lang, lang)} 번역")
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 데이터 입력 및 번역
            total_elements = len(text_elements)
            
            for idx, element in enumerate(text_elements, 1):
                if verbose:
                    print(f"처리 중: {idx}/{total_elements} - {element['text'][:50]}...")
                
                row = idx + 1
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=element['type'])
                ws.cell(row=row, column=3, value=element['tag'])
                ws.cell(row=row, column=4, value=element['text'])
                
                col = 5
                
                # 선택된 언어로 번역
                for lang in languages:
                    translated_text = self.translate_text(element['text'], lang, verbose)
                    ws.cell(row=row, column=col, value=translated_text)
                    col += 1
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(file_path)
            
            if verbose:
                print(f"엑셀 파일이 저장되었습니다: {file_path}")
            
            return True
            
        except Exception as e:
            print(f"엑셀 파일 생성 오류: {str(e)}")
            return False
    
    def process_url(self, url, output_file=None, languages=['en', 'zh-cn', 'vi'], verbose=True):
        """URL 처리 메인 함수"""
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        
        if not output_file:
            output_file = f"웹텍스트_추출_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 텍스트 추출
        text_elements = self.extract_text_from_url(url, verbose)
        
        if not text_elements:
            print("텍스트 추출에 실패했습니다.")
            return False
        
        # 엑셀 파일 생성
        success = self.create_excel_file(text_elements, output_file, languages, verbose)
        
        if success:
            print(f"✅ 작업 완료! 파일: {output_file}")
            return True
        else:
            print("❌ 작업 실패")
            return False

def main():
    parser = argparse.ArgumentParser(
        description="웹 텍스트 추출 및 번역 자동화 도구",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
사용 예시:
  python cli_extractor.py https://example.com
  python cli_extractor.py https://example.com -o result.xlsx
  python cli_extractor.py https://example.com -l en zh-cn vi ja
  python cli_extractor.py https://example.com --quiet
        """
    )
    
    parser.add_argument('url', help='추출할 웹페이지 URL')
    parser.add_argument('-o', '--output', help='출력 엑셀 파일명')
    parser.add_argument('-l', '--languages', nargs='+', 
                       default=['en', 'zh-cn', 'vi'],
                       help='번역할 언어 코드 (기본값: en zh-cn vi)')
    parser.add_argument('-q', '--quiet', action='store_true',
                       help='자세한 출력 비활성화')
    parser.add_argument('--list-languages', action='store_true',
                       help='지원되는 언어 코드 목록 표시')
    
    args = parser.parse_args()
    
    if args.list_languages:
        print("지원되는 언어 코드:")
        lang_codes = {
            'en': '영어 (English)',
            'zh-cn': '중국어 간체 (简体中文)',
            'zh-tw': '중국어 번체 (繁體中文)',
            'vi': '베트남어 (Tiếng Việt)',
            'ja': '일본어 (日本語)',
            'es': '스페인어 (Español)',
            'fr': '프랑스어 (Français)',
            'de': '독일어 (Deutsch)',
            'ru': '러시아어 (Русский)',
            'pt': '포르투갈어 (Português)',
            'it': '이탈리아어 (Italiano)',
            'ar': '아랍어 (العربية)',
            'hi': '힌디어 (हिन्दी)',
            'th': '태국어 (ไทย)'
        }
        for code, name in lang_codes.items():
            print(f"  {code}: {name}")
        return
    
    extractor = CLIWebTextExtractor()
    
    try:
        success = extractor.process_url(
            args.url, 
            args.output, 
            args.languages, 
            not args.quiet
        )
        
        sys.exit(0 if success else 1)
        
    except KeyboardInterrupt:
        print("\n작업이 사용자에 의해 중단되었습니다.")
        sys.exit(1)
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main() 