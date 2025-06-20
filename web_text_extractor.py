import requests
from bs4 import BeautifulSoup, Comment
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from googletrans import Translator
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import re
from urllib.parse import urljoin, urlparse
import os
from datetime import datetime

class WebTextExtractor:
    def __init__(self):
        self.translator = Translator()
        self.setup_gui()
        
    def setup_gui(self):
        """GUI 인터페이스 설정"""
        self.root = tk.Tk()
        self.root.title("웹 텍스트 추출 및 번역 자동화 도구")
        self.root.geometry("800x600")
        self.root.configure(bg='#f0f0f0')
        
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목
        title_label = ttk.Label(main_frame, text="웹 텍스트 추출 및 번역 도구", 
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # URL 입력 섹션
        url_frame = ttk.LabelFrame(main_frame, text="URL 설정", padding="10")
        url_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(url_frame, text="웹사이트 URL (여러 개인 경우 줄바꿈으로 구분):").pack(anchor=tk.W)
        
        # URL 입력을 위한 텍스트 영역
        url_text_frame = ttk.Frame(url_frame)
        url_text_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.url_text = tk.Text(url_text_frame, height=4, wrap=tk.WORD)
        url_scrollbar = ttk.Scrollbar(url_text_frame, orient=tk.VERTICAL, command=self.url_text.yview)
        self.url_text.configure(yscrollcommand=url_scrollbar.set)
        
        self.url_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        url_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 번역 언어 선택
        lang_frame = ttk.LabelFrame(main_frame, text="번역 언어 선택", padding="10")
        lang_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.translate_english = tk.BooleanVar(value=True)
        self.translate_chinese = tk.BooleanVar(value=True)
        self.translate_vietnamese = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(lang_frame, text="영어 (English)", 
                       variable=self.translate_english).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Checkbutton(lang_frame, text="중국어 (简体中文)", 
                       variable=self.translate_chinese).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Checkbutton(lang_frame, text="베트남어 (Tiếng Việt)", 
                       variable=self.translate_vietnamese).pack(side=tk.LEFT)
        
        # 출력 파일 설정
        file_frame = ttk.LabelFrame(main_frame, text="출력 파일 설정", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        file_path_frame = ttk.Frame(file_frame)
        file_path_frame.pack(fill=tk.X)
        
        ttk.Label(file_path_frame, text="엑셀 파일 경로:").pack(anchor=tk.W)
        path_frame = ttk.Frame(file_path_frame)
        path_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.file_path_var = tk.StringVar()
        self.file_path_var.set(f"웹텍스트_추출_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        ttk.Entry(path_frame, textvariable=self.file_path_var, width=60).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(path_frame, text="찾아보기", command=self.browse_file).pack(side=tk.RIGHT, padx=(10, 0))
        
        # 실행 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.extract_button = ttk.Button(button_frame, text="텍스트 추출 및 번역 시작", 
                                        command=self.start_extraction, style='Accent.TButton')
        self.extract_button.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(button_frame, text="종료", command=self.root.quit).pack(side=tk.RIGHT)
        
        # 진행 상황
        progress_frame = ttk.LabelFrame(main_frame, text="진행 상황", padding="10")
        progress_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        self.progress_var = tk.StringVar(value="대기 중...")
        ttk.Label(progress_frame, textvariable=self.progress_var).pack(anchor=tk.W)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress_bar.pack(fill=tk.X, pady=(10, 0))
        
        # 로그 텍스트 영역
        log_frame = ttk.Frame(progress_frame)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        self.log_text = tk.Text(log_frame, height=10, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
    def browse_file(self):
        """파일 경로 선택 다이얼로그"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.file_path_var.set(filename)
    
    def log_message(self, message):
        """로그 메시지 추가"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def extract_text_from_url(self, url):
        """웹페이지에서 텍스트 추출"""
        try:
            self.log_message(f"웹페이지 접속 중: {url}")
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            response.encoding = response.apparent_encoding
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # 불필요한 태그 및 주석 제거
            for tag in soup(['script', 'style', 'nav', 'header', 'footer', 'aside']):
                tag.decompose()
                
            # HTML 주석 제거
            for comment in soup.find_all(string=lambda text: isinstance(text, Comment)):
                comment.extract()
            
            # 텍스트 추출 - DOM 순서대로 개별 요소별로
            text_elements = []
            seen_texts = set()  # 중복 방지용
            
            # 메인 컨텐츠 영역 찾기
            main_content = soup.find('main') or soup.find('article') or soup.find(class_=re.compile(r'content|main|body', re.I)) or soup.find('body')
            
            if main_content:
                # DOM을 순회하면서 개별 텍스트 요소 추출
                self._extract_text_recursively(main_content, text_elements, seen_texts)
            else:
                # main_content를 찾지 못한 경우 전체 body에서 추출
                body = soup.find('body')
                if body:
                    self._extract_text_recursively(body, text_elements, seen_texts)
            
            self.log_message(f"총 {len(text_elements)}개의 텍스트 요소를 순차적으로 추출했습니다.")
            return text_elements
            
        except requests.RequestException as e:
            self.log_message(f"웹페이지 접속 오류: {str(e)}")
            return []
        except Exception as e:
            self.log_message(f"텍스트 추출 오류: {str(e)}")
            return []
    
    def _extract_text_recursively(self, element, text_elements, seen_texts):
        """재귀적으로 텍스트 추출"""
        for child in element.children:
            # 텍스트 노드인 경우 (순수 텍스트) - 주석 제외
            if child.name is None:
                # HTML 주석인지 확인
                if isinstance(child, Comment):
                    continue
                    
                text = str(child).strip()
                if text and len(text) > 2 and not text.startswith('<'):
                    text_clean = text.replace('\n', ' ').replace('\t', ' ').strip()
                    # 공백만 있거나 특수문자만 있는 경우 제외
                    if text_clean and text_clean not in seen_texts and len(text_clean.replace(' ', '')) > 1:
                        text_elements.append({
                            'type': 'content',
                            'tag': 'text',
                            'text': text_clean
                        })
                        seen_texts.add(text_clean)
            
            # HTML 요소인 경우
            elif child.name:
                # 제목 태그들은 개별적으로 처리
                if child.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                    text = child.get_text(strip=True)
                    if text and len(text) > 1:
                        text_clean = text.replace('\n', ' ').replace('\t', ' ').strip()
                        if text_clean and text_clean not in seen_texts:
                            text_elements.append({
                                'type': 'heading',
                                'tag': child.name,
                                'text': text_clean
                            })
                            seen_texts.add(text_clean)
                
                # 단락, 리스트 항목 등은 개별적으로 처리
                elif child.name in ['p', 'li', 'td', 'th', 'blockquote', 'pre']:
                    text = child.get_text(strip=True)
                    if text and len(text) > 2:
                        text_clean = text.replace('\n', ' ').replace('\t', ' ').strip()
                        # 중복 체크
                        is_duplicate = False
                        for seen_text in seen_texts:
                            if text_clean == seen_text:
                                is_duplicate = True
                                break
                            # 포함 관계 체크 (90% 이상 겹치면 중복으로 간주)
                            if text_clean in seen_text and len(text_clean) > len(seen_text) * 0.9:
                                is_duplicate = True
                                break
                            if seen_text in text_clean and len(seen_text) > len(text_clean) * 0.9:
                                # 더 긴 텍스트로 교체
                                text_elements[:] = [elem for elem in text_elements if elem['text'] != seen_text]
                                seen_texts.discard(seen_text)
                                break
                        
                        if not is_duplicate and text_clean:
                            text_elements.append({
                                'type': 'content',
                                'tag': child.name,
                                'text': text_clean
                            })
                            seen_texts.add(text_clean)
                
                # 인라인 요소들 - 텍스트가 의미있는 경우만
                elif child.name in ['span', 'a', 'strong', 'b', 'em', 'i', 'code', 'label']:
                    text = child.get_text(strip=True)
                    if text and len(text) > 1:
                        text_clean = text.replace('\n', ' ').replace('\t', ' ').strip()
                        if text_clean and text_clean not in seen_texts:
                            # 너무 짧거나 의미없는 텍스트 제외
                            if len(text_clean) > 2 and not text_clean.isdigit():
                                text_elements.append({
                                    'type': 'content',
                                    'tag': child.name,
                                    'text': text_clean
                                })
                                seen_texts.add(text_clean)
                
                # div, section 등 컨테이너 요소는 재귀적으로 처리
                elif child.name in ['div', 'section', 'article', 'ul', 'ol', 'table', 'tbody', 'thead', 'tr']:
                    # 하위 요소들을 재귀적으로 처리
                    self._extract_text_recursively(child, text_elements, seen_texts)
    
    def translate_text(self, text, target_lang):
        """텍스트 번역"""
        try:
            # 번역할 텍스트가 너무 길면 분할
            if len(text) > 4000:
                # 문장 단위로 분할
                sentences = re.split(r'[.!?。！？]', text)
                translated_sentences = []
                
                current_chunk = ""
                for sentence in sentences:
                    if len(current_chunk + sentence) < 4000:
                        current_chunk += sentence + "."
                    else:
                        if current_chunk:
                            result = self.translator.translate(current_chunk, dest=target_lang)
                            translated_sentences.append(result.text)
                        current_chunk = sentence + "."
                
                if current_chunk:
                    result = self.translator.translate(current_chunk, dest=target_lang)
                    translated_sentences.append(result.text)
                
                return " ".join(translated_sentences)
            else:
                result = self.translator.translate(text, dest=target_lang)
                return result.text
                
        except Exception as e:
            self.log_message(f"번역 오류 ({target_lang}): {str(e)}")
            return f"[번역 실패: {text[:50]}...]"
    
    def create_excel_file(self, text_elements, file_path):
        """엑셀 파일 생성"""
        try:
            self.log_message("엑셀 파일 생성 중...")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "웹 텍스트 추출 결과"
            
            # 헤더 설정
            headers = ["번호", "URL", "유형", "태그", "원본 텍스트(한국어)"]
            
            if self.translate_english.get():
                headers.append("영어 번역")
            if self.translate_chinese.get():
                headers.append("중국어 번역")
            if self.translate_vietnamese.get():
                headers.append("베트남어 번역")
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 데이터 입력 및 번역
            total_elements = len(text_elements)
            
            for idx, element in enumerate(text_elements, 1):
                self.progress_var.set(f"처리 중... ({idx}/{total_elements})")
                self.log_message(f"처리 중: {idx}/{total_elements} - {element['text'][:50]}...")
                
                row = idx + 1
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=element.get('url', ''))
                ws.cell(row=row, column=3, value=element['type'])
                ws.cell(row=row, column=4, value=element['tag'])
                ws.cell(row=row, column=5, value=element['text'])
                
                col = 6
                
                # 영어 번역
                if self.translate_english.get():
                    english_text = self.translate_text(element['text'], 'en')
                    ws.cell(row=row, column=col, value=english_text)
                    col += 1
                
                # 중국어 번역
                if self.translate_chinese.get():
                    chinese_text = self.translate_text(element['text'], 'zh-cn')
                    ws.cell(row=row, column=col, value=chinese_text)
                    col += 1
                
                # 베트남어 번역
                if self.translate_vietnamese.get():
                    vietnamese_text = self.translate_text(element['text'], 'vi')
                    ws.cell(row=row, column=col, value=vietnamese_text)
                    col += 1
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(file_path)
            self.log_message(f"엑셀 파일이 저장되었습니다: {file_path}")
            
            return True
            
        except Exception as e:
            self.log_message(f"엑셀 파일 생성 오류: {str(e)}")
            return False
    
    def start_extraction(self):
        """텍스트 추출 및 번역 시작"""
        urls_text = self.url_text.get("1.0", tk.END).strip()
        if not urls_text:
            messagebox.showerror("오류", "URL을 입력해주세요.")
            return
        
        # URL 목록 정리
        urls = []
        for line in urls_text.splitlines():
            url = line.strip()
            if url:
                if not url.startswith(('http://', 'https://')):
                    url = 'https://' + url
                urls.append(url)
        
        if not urls:
            messagebox.showerror("오류", "유효한 URL을 입력해주세요.")
            return
        
        file_path = self.file_path_var.get().strip()
        if not file_path:
            messagebox.showerror("오류", "출력 파일 경로를 설정해주세요.")
            return
        
        # 별도 스레드에서 실행
        self.extract_button.config(state='disabled')
        self.progress_bar.start()
        
        thread = threading.Thread(target=self.extract_and_translate, args=(urls, file_path))
        thread.daemon = True
        thread.start()
    
    def extract_and_translate(self, urls, file_path):
        """실제 추출 및 번역 작업 수행"""
        try:
            all_results = []
            total_urls = len(urls)
            
            # 각 URL별로 텍스트 추출
            for idx, url in enumerate(urls, 1):
                self.progress_var.set(f"URL 처리 중... ({idx}/{total_urls}): {url}")
                self.log_message(f"=== URL {idx}/{total_urls} 처리 시작: {url} ===")
                
                text_elements = self.extract_text_from_url(url)
                
                if text_elements:
                    # URL 정보를 각 텍스트 요소에 추가
                    for element in text_elements:
                        element['url'] = url
                        element['url_index'] = idx
                    all_results.extend(text_elements)
                    self.log_message(f"URL {idx} 처리 완료: {len(text_elements)}개 텍스트 추출")
                else:
                    self.log_message(f"URL {idx} 처리 실패: 텍스트를 추출할 수 없습니다.")
            
            if not all_results:
                self.progress_var.set("모든 URL에서 텍스트 추출 실패")
                messagebox.showerror("오류", "모든 웹페이지에서 텍스트를 추출할 수 없습니다.")
                return
            
            # 엑셀 파일 생성
            success = self.create_excel_file(all_results, file_path)
            
            if success:
                self.progress_var.set("작업 완료!")
                messagebox.showinfo("완료", f"작업이 완료되었습니다!\n처리된 URL: {total_urls}개\n추출된 텍스트: {len(all_results)}개\n파일 경로: {file_path}")
            else:
                self.progress_var.set("작업 실패")
                messagebox.showerror("오류", "엑셀 파일 생성 중 오류가 발생했습니다.")
                
        except Exception as e:
            self.progress_var.set("오류 발생")
            self.log_message(f"전체 작업 오류: {str(e)}")
            messagebox.showerror("오류", f"작업 중 오류가 발생했습니다: {str(e)}")
        
        finally:
            self.extract_button.config(state='normal')
            self.progress_bar.stop()
    
    def run(self):
        """GUI 실행"""
        self.root.mainloop()

if __name__ == "__main__":
    app = WebTextExtractor()
    app.run() 